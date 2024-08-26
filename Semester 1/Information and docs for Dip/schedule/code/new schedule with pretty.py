import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Creating the data for the schedule
data = {
    "Time": [
        "6:00-7:00 AM", "7:00-7:30 AM", "7:30-8:30 AM", "8:30-9:30 AM",
        "9:00-10:00 AM", "10:00-11:00 AM", "11:00-12:00 PM", "12:00-1:00 PM",
        "1:00-2:00 PM", "2:00-3:00 PM", "3:00-4:00 PM", "4:00-5:00 PM",
        "5:00-6:00 PM", "6:00-7:00 PM", "7:00-8:00 PM", "8:00-9:00 PM",
        "9:00-10:00 PM", "10:00-11:00 PM", "11:00-12:00 AM"
    ],
    "Monday": [
        "-", "-", "-", "-",
        "Wake up, Breakfast", "Study", "Study", "Study",
        "Lunch/Study Break", "Study", "Study", "Study",
        "Dinner/Leisure", "Study", "Study", "Leisure",
        "Leisure", "Rest/Free Time", "Rest"
    ],
    "Tuesday": [
        "Travel to class", "Travel to class", "Class (8:30 AM-3:30 PM)", "Class",
        "Class", "Class", "Class", "Class",
        "Class", "Travel/Rest", "Study", "Study",
        "Dinner", "Study", "Study", "Leisure",
        "Leisure", "Rest/Free Time", "Rest"
    ],
    "Wednesday": [
        "Travel to class", "Travel to class", "Class (8:30 AM-5:30 PM)", "Class",
        "Class", "Class", "Class", "Class",
        "Class", "Travel/Rest", "Study/Rest", "Study/Rest",
        "Dinner", "Study", "Study", "Leisure",
        "Leisure", "Rest/Free Time", "Rest"
    ],
    "Thursday": [
        "-", "-", "-", "-",
        "Wake up, Breakfast", "Study", "Study", "Study",
        "Lunch/Study Break", "Study", "Study", "Study",
        "Dinner", "Work (2:00 PM-12:00 AM)", "Work", "Work",
        "Work", "Work", "Work"
    ],
    "Friday": [
        "-", "-", "-", "-",
        "Wake up, Breakfast", "Study", "Study", "Study",
        "Lunch/Study Break", "Study", "Study", "Study",
        "Dinner", "Work (2:00 PM-12:00 AM)", "Work", "Work",
        "Work", "Work", "Work"
    ],
    "Saturday": [
        "-", "-", "-", "-",
        "Wake up, Breakfast", "Workout", "Workout", "Workout",
        "Lunch/Leisure", "Study", "Study", "Study",
        "Dinner", "Work (2:00 PM-12:00 AM)", "Work", "Work",
        "Work", "Work", "Work"
    ],
    "Sunday": [
        "-", "-", "-", "-",
        "Wake up, Breakfast", "Leisure", "Leisure", "Leisure",
        "Leisure", "Leisure", "Leisure", "Leisure",
        "Dinner", "Leisure", "Leisure", "Leisure",
        "Leisure", "Rest", "Rest"
    ]
}

# Create a DataFrame
df = pd.DataFrame(data)

# Create a new Excel workbook and add a sheet
wb = Workbook()
ws = wb.active
ws.title = "Weekly Schedule"

# Set the title font
title_font = Font(size=14, bold=True)
header_font = Font(size=12, bold=True)
align_center = Alignment(horizontal="center", vertical="center")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Add headers
headers = ["Time", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
ws.append(headers)

for cell in ws[1]:  # Iterate directly over the first row cells
    cell.font = header_font
    cell.alignment = align_center
    cell.border = thin_border

# Add the data from the DataFrame
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
    ws.append(row)
    for c_idx, cell in enumerate(ws[r_idx], 1):
        cell.alignment = align_center
        cell.border = thin_border
        if c_idx > 1:
            cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")

# Add a column for checkboxes
for col in range(2, 9):  # Monday to Sunday
    ws.cell(row=1, column=col).value = headers[col-1]
    for row in range(2, len(df) + 2):
        ws.cell(row=row, column=col).value = ""  # Placeholder for marking off tasks

# Add daily and weekly score rows
ws.append(["Daily Score"] + [""] * 7)
ws.append(["Weekly Score"] + [""] * 7)

# Format the "Daily Score" and "Weekly Score" rows
for cell in ws[ws.max_row-1]:  # Formatting the "Daily Score" row
    cell.font = header_font
    cell.alignment = align_center
    cell.border = thin_border

for cell in ws[ws.max_row]:  # Formatting the "Weekly Score" row
    cell.font = header_font
    cell.alignment = align_center
    cell.border = thin_border

# Auto-adjust the column widths for better readability
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Save the workbook
wb.save("C:\\Users\\NickS\\Documents\\formatted_weekly_schedule.xlsx")

